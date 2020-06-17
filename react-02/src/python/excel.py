from openpyxl import load_workbook
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.utils import quote_sheetname

# def data_validation():
#     wb = load_workbook('excel.xlsx')
#     ws = wb['product']
#     print(ws)
#     dvDecimal = DataValidation(type="decimal")
#     dvDecimal.add('C2:C10')
#     ws.add_data_validation(dvDecimal)
#     print(dvDecimal.error)

# data_validation()

def breakIntoDict():
    wb = load_workbook('excel.xlsx')
    builtDict = {}
    for sheet in wb.sheetnames:
        builtDict[sheet] = {}
        rowCount = 1
        for row in wb[sheet].iter_rows(min_row=2,min_col=1):
            if row[0].value is not None:
                col = 0
                builtDict[sheet][rowCount] = {}
                for cell in row:
                    builtDict[sheet][rowCount][wb[sheet][1][col].value] = cell.value
                    col += 1
            rowCount += 1
    # print(builtDict)
    return builtDict

# returns the customer ID(from invoice dictionary)
def getCustID(invoice) :
    for inv in xlDict['invoices']:
        if invoice == xlDict['invoices'][inv]['Invoice']:
            return xlDict['invoices'][inv]['Customer']

# one = dictionary, two = dictionary, same = dict one column, match = dict two column, where = dictionary two column, estar = match for where
def innerjoin(one,two,same,match,where=None,estar=None):
    outDict = {}
    for oneRow in one:
        for twoRow in two:
            if one[oneRow][same] == two[twoRow][match]:
                try:
                    theTest = (two[twoRow][where] == estar)
                except:
                    theTest = False
                if (theTest) or where==None:
                    index = len(outDict) +1
                    outDict[index] = {}
                    for item in one[oneRow]:
                            # print(item,one[oneRow][item])
                            outDict[index][item] = one[oneRow][item]
                    for element in two[twoRow]:
                            # print(element,two[twoRow][element])
                            outDict[index][element] = two[twoRow][element]
    return outDict


def printInvoice(invObj):
    invoiceFile = 'Invoice'+str(invObj[1]['Invoice'])
    with open(invoiceFile+'.txt','w') as invoice:
        # report.write(b[0]+'-'+b[1]+':'+str(tupleDict[b])+'\n')
        invoice.write('Invoice: '+str(invObj[1]['Invoice'])+'\n')
        invoice.write('Date: '+str(invObj[1]['Date'].date())+'\n')
        invoice.write('Customer: '+invObj[1]['Name']+'\n')
        invoice.write('Customer ID: '+str(invObj[1]['Customer'])+'\n\n')
        invoice.write("{: <5} {: <31} {: <8}\n".format('Qty:','Product:','Price:'))
        invoice.write("{: <5} {: <31} {: <8}\n".format('====','========','======'))
        total = 0
        for row in invObj:
            qty = str(invObj[row]['Quantity'])
            product = invObj[row]['Beer Name']
            total+=invObj[row]['Price']
            price = str(invObj[row]['Price'])
            invoice.write("{: <5} {: <31} {: <8}\n".format(qty,product,price))
        invoice.write("{: >44}\n".format('TOTAL:'))
        total = round(total,2)
        invoice.write("{: >44}\n".format(total))

xlDict = breakIntoDict()

# This block of code creates a list of existing invoice
# number to display to the user
invoices = xlDict['invoices']
invoiceList = []
for invoice in invoices:
    invoiceList.append(invoices[invoice]['Invoice'])
print(invoiceList)

# This code gets a valid invoice number from the user
invoiceReq = 0
if __name__ == "__main__":
    while invoiceReq not in invoiceList:
        invoiceReq = int(input('Enter existing invoice:'))
    first = innerjoin(xlDict['customers'],xlDict['invoices'],'ID','Customer','Invoice',invoiceReq)
    second = innerjoin(first,xlDict['invoice_line_items'],'Invoice','Invoice','Invoice',invoiceReq)
    third = innerjoin(second,xlDict['product'],'Product','ID')
    print(third)
    printInvoice(third)


